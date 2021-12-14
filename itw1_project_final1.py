#attendance management system
import pandas as pd
import xlrd
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import xlsxwriter
#location of sheet
loc = "attendance_record.xlsx"
#maintaing the current  numbers
current_row = 3
current_column =3 
current_sheet_name =''
#
#required variables for[]
#list of all courses
total_course = ['OOPS','CSO','MI']


#creating excel_sheet
def create_new_excelfile():
 #   writer = pd.ExcelWriter(loc, engine='xlsxwriter')
 #   writer.save()
     workbook = xlsxwriter.Workbook()
     worksheet = workbook.add_worksheet('OOPS')
     worksheet.set_column('A:A',20)
     worksheet = workbook.add_worksheet('CSO')
     worksheet.set_column('A:A',20)
     worksheet = workbook.add_worksheet('MI')
     worksheet.set_column('A:A',20)

def add_students():
    df = pd.DataFrame({'Roll No': ['BT19CSE001','BT19CSE002','BT19CSE003','BT19CSE004','BT19CSE005','BT19CSE006'],
        'Name': ['Pranav','Kirteeraj','Harsh','Ekta','Diwyani','shyam'],
        'Total':[0,0,0,0,0,0]
        })
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(loc, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    for course in total_course:
        df.to_excel(writer, sheet_name=course, index=False)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()



# func for displaying the all courses
def display_courses(): 
    print("Select the index number for choosing the course\n")   
    for i in range(len(total_course)):
        print(str(i+1) + ". " + total_course[i])

#c
def display_students():
    #for printing the course list
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    wr = openpyxl.load_workbook(filename=loc)
    sheet = wr[current_sheet_name]
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
    for i in range(1, sheet.max_row + 1): 
        cell_obj = sheet.cell(row = i, column = 2) 
        print("|",cell_obj.value)
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
#c
def display_record():
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    print(" displaying the record of all student till current date\n")
    df = pd.read_excel(loc,sheet_name = current_sheet_name)
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
    print(df)
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
    return

def mark_attendance():
    #to disply course list
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    wb = openpyxl.load_workbook(filename = loc)
    sheet = wb.active
    sheet = wb[current_sheet_name]
    print("p for present\na for absent\n")
    current_column  = sheet.max_column+1
    sheet.cell(row = 1,column = current_column).value = ("lec "+str(sheet.max_column-2))

    for i in range(2,sheet.max_row+1):
        att = input(sheet.cell(row=i,column = 2).value + " : ")
    
       # if(att == 1):
        sheet.cell(row=i,column = current_column).value=att
     #  elif(att == 0):
            #sheet.cell(row=i,column = current_column).value="a"
        wb.save(loc)
     #
    for r in range(1,sheet.max_row +1):
        count = 0
        for c in range(4, sheet.max_column + 1): 
            cell_obj = sheet.cell(row = r, column = c).value
            if(cell_obj == "p"):
                count+=1
            sheet.cell(row = r,column = 3).value = count
            wb.save(loc)
    
                
def eligibility_for_exam():
    #open sheet
    count =0
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    wr = openpyxl.load_workbook(loc)
    sheet = wr[current_sheet_name]
    #printing for all
    t = int(input("1. Select to check eligibility of all the students\n2. Select to check eligibility of a particular student by name\n"))
    if(t == 1):
        for r in range(1,sheet.max_row+1):
            temp = sheet.cell(row = r,column = 3).value
            if(temp/40 >= 0.75):
                print(sheet.cell(row = r,column = 2).value , " - Eligible\n")
            else:
                print(sheet.cell(row = r,column = 2).value , " - Not Eligible\n")
    elif(t == 2):
        n = input("Enter the name of student \n")
        for r in range(1,sheet.max_row+1):
            if(n == sheet.cell(row = r,column = 2).value):
                temp = sheet.cell(row = r,column = 3).value
                break
            #else:
             #   print("Please enter the valid name of a student\n")
        if(temp/40 >= 0.75):
            print(sheet.cell(row = r,column = 2).value , " - Eligible\n")
        else:
            print(sheet.cell(row = r,column = 2).value , " - Not Eligible\n")
                
                
   
def main():
    #create_new_excelfile()
    #add_students()
    while(1):
        print("1. Enter to Display the registered students for particular course\n")
        print("2. Enter to display the attendace record for registerd student\n")
        print("3. Enter to mark the attendance record of students\n")
        print("4. Enter to display if the student is eligible for exam or not\n")
        print("0. To Exit")
        switch_key = int(input())
        #switch
        if(switch_key == 1):
            display_students()
        elif(switch_key == 2):
            display_record()
        elif(switch_key == 3):
            mark_attendance()
        elif(switch_key == 4):
            eligibility_for_exam()
        elif(switch_key == 0):
            break
    
        else:
            print("Enter the valid choice\n")

    return

if __name__ == "__main__":
    main()
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import xlsxwriter
#location of sheet
loc = "attendance_record.xlsx"
#maintaing the current  numbers
current_row = 3
current_column =3 
current_sheet_name =''
#
#required variables for[]
#list of all courses
total_course = ['OOPS','CSO','MI']


#creating excel_sheet
def create_new_excelfile():
 #   writer = pd.ExcelWriter(loc, engine='xlsxwriter')
 #   writer.save()
     workbook = xlsxwriter.Workbook()
     worksheet = workbook.add_worksheet('OOPS')
     worksheet.set_column('A:A',20)
     worksheet = workbook.add_worksheet('CSO')
     worksheet.set_column('A:A',20)
     worksheet = workbook.add_worksheet('MI')
     worksheet.set_column('A:A',20)

def add_students():
    df = pd.DataFrame({'Roll No': ['BT19CSE001','BT19CSE002','BT19CSE003','BT19CSE004','BT19CSE005','BT19CSE006'],
        'Name': ['Pranav','Kirteeraj','Harsh','Ekta','Diwyani','shyam'],
        'Total':[0,0,0,0,0,0]
        })
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(loc, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    for course in total_course:
        df.to_excel(writer, sheet_name=course, index=False)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()



# func for displaying the all courses
def display_courses(): 
    print("Select the index number for choosing the course\n")   
    for i in range(len(total_course)):
        print(str(i+1) + ". " + total_course[i])

#c
def display_students():
    #for printing the course list
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    wr = openpyxl.load_workbook(filename=loc)
    sheet = wr[current_sheet_name]
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
    for i in range(1, sheet.max_row + 1): 
        cell_obj = sheet.cell(row = i, column = 2) 
        print("|",cell_obj.value)
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
#c
def display_record():
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    print(" displaying the record of all student till current date\n")
    df = pd.read_excel(loc,sheet_name = current_sheet_name)
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
    print(df)
    print("----------------------------------------------------------------------------------------------------------------------------------------\n")
    return

def mark_attendance():
    #to disply course list
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    wb = openpyxl.load_workbook(filename = loc)
    sheet = wb.active
    sheet = wb[current_sheet_name]
    print("p for present\na for absent\n")
    current_column  = sheet.max_column+1
    sheet.cell(row = 1,column = current_column).value = ("lec "+str(sheet.max_column-2))

    for i in range(2,sheet.max_row+1):
        att = input(sheet.cell(row=i,column = 2).value + " : ")
    
       # if(att == 1):
        sheet.cell(row=i,column = current_column).value=att
     #  elif(att == 0):
            #sheet.cell(row=i,column = current_column).value="a"
        wb.save(loc)
     #
    for r in range(1,sheet.max_row +1):
        count = 0
        for c in range(4, sheet.max_column + 1): 
            cell_obj = sheet.cell(row = r, column = c).value
            if(cell_obj == "p"):
                count+=1
            sheet.cell(row = r,column = 3).value = count
            wb.save(loc)
    
                
def eligibility_for_exam():
    #open sheet
    count =0
    display_courses()
    temp = int(input())
    current_sheet_name = total_course[temp-1]
    wr = openpyxl.load_workbook(loc)
    sheet = wr[current_sheet_name]
    #printing for all
    t = int(input("1. Select to check eligibility of all the students\n2. Select to check eligibility of a particular student by name\n"))
    if(t == 1):
        for r in range(1,sheet.max_row+1):
            temp = sheet.cell(row = r,column = 3).value
            if(temp/40 >= 0.75):
                print(sheet.cell(row = r,column = 2).value , " - Eligible\n")
            else:
                print(sheet.cell(row = r,column = 2).value , " - Not Eligible\n")
    elif(t == 2):
        n = input("Enter the name of student \n")
        for r in range(1,sheet.max_row+1):
            if(n == sheet.cell(row = r,column = 2).value):
                temp = sheet.cell(row = r,column = 3).value
                break
            
        if(temp/40 >= 0.75):
            print(sheet.cell(row = r,column = 2).value , " - Eligible\n")
        else:
            print(sheet.cell(row = r,column = 2).value , " - Not Eligible\n")
                
                
  

def main():
    #create_new_excelfile()
    #add_students()
    while(1):
        print("1. Enter to Display the registered students for particular course\n")
        print("2. Enter to display the attendace record for registerd student\n")
        print("3. Enter to mark the attendance record of students\n")
        print("4. Enter to display if the student is eligible for exam or not\n")
        print("0. To Exit")
        switch_key = int(input())
        #switch
        if(switch_key == 1):
            display_students()
        elif(switch_key == 2):
            display_record()
        elif(switch_key == 3):
            mark_attendance()
        elif(switch_key == 4):
            eligibility_for_exam()
        elif(switch_key == 0):
            break
            return
        else:
            print("Enter the valid choice\n")

    return

if __name__ == "__main__":
    main()